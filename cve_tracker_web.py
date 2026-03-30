#!/usr/bin/env python3
"""
CVE Tracker — Web Frontend
Run: python3 cve_tracker_web.py
Then open: http://localhost:5001
"""

import os
import sqlite3
from datetime import datetime
from flask import Flask, render_template_string, request, redirect, url_for, jsonify

DB_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'cve_tracker.db')

app = Flask(__name__)

SEVERITY_ORDER = {'Critical': 0, 'High': 1, 'Medium': 2, 'Moderate': 2, 'Low': 3, 'Informational': 4}
SEVERITY_COLORS = {
    'Critical':      '#ef4444',
    'High':          '#f97316',
    'Medium':        '#eab308',
    'Moderate':      '#eab308',
    'Low':           '#22c55e',
    'Informational': '#64748b',
}
STATUS_COLORS = {
    'open':           '#ef4444',
    'investigating':  '#f97316',
    'accepted':       '#a855f7',
    'resolved':       '#22c55e',
    'false_positive': '#64748b',
}
VALID_STATUSES = ['open', 'investigating', 'accepted', 'resolved', 'false_positive']

ENABLED_MODES = {
    'not_enabled': 'Not enabled by default in any deployment',
    'block_notify': ['Security-Optimized Block / Notify', 'Default Block / Notify'],
    'all': None,
}

def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def dict_row(row):
    return dict(row) if row else None

HTML = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>CVE Tracker</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600;700&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
:root {
  --bg: #050810;
  --surface: #0c1120;
  --surface2: #111827;
  --border: #1e293b;
  --text: #e2e8f0;
  --muted: #475569;
  --mono: 'IBM Plex Mono', monospace;
  --sans: 'IBM Plex Sans', sans-serif;
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body { background: var(--bg); color: var(--text); font-family: var(--sans); min-height: 100vh; }

/* ── Layout ── */
.shell { display: flex; min-height: 100vh; }
.sidebar {
  width: 220px; flex-shrink: 0;
  background: var(--surface);
  border-right: 1px solid var(--border);
  padding: 24px 0;
  position: fixed; top: 0; left: 0; bottom: 0;
  display: flex; flex-direction: column;
}
.main { margin-left: 220px; padding: 32px; min-height: 100vh; }

/* ── Sidebar ── */
.logo {
  font-family: var(--mono); font-size: 11px; font-weight: 700;
  letter-spacing: 3px; color: #ef4444; padding: 0 20px 24px;
  border-bottom: 1px solid var(--border); margin-bottom: 20px;
  line-height: 1.6;
}
.logo span { color: var(--muted); display: block; font-size: 9px; letter-spacing: 2px; margin-top: 2px; }
.nav-label { font-family: var(--mono); font-size: 9px; color: var(--muted); letter-spacing: 2px; text-transform: uppercase; padding: 0 20px; margin-bottom: 6px; }
.nav-item { display: flex; align-items: center; gap: 10px; padding: 9px 20px; font-size: 13px; color: var(--muted); cursor: pointer; transition: .15s; text-decoration: none; border-left: 2px solid transparent; }
.nav-item:hover { color: var(--text); background: rgba(255,255,255,.03); }
.nav-item.active { color: var(--text); border-left-color: #ef4444; background: rgba(239,68,68,.06); }
.nav-count { margin-left: auto; font-family: var(--mono); font-size: 10px; background: var(--surface2); border: 1px solid var(--border); padding: 1px 6px; border-radius: 3px; }
.sidebar-footer { margin-top: auto; padding: 16px 20px; border-top: 1px solid var(--border); font-family: var(--mono); font-size: 9px; color: var(--muted); }

/* ── Stats bar ── */
.stats-bar { display: flex; gap: 12px; margin-bottom: 28px; flex-wrap: wrap; }
.stat { background: var(--surface); border: 1px solid var(--border); border-radius: 8px; padding: 14px 18px; min-width: 100px; }
.stat-num { font-family: var(--mono); font-size: 24px; font-weight: 700; line-height: 1; }
.stat-label { font-size: 10px; color: var(--muted); letter-spacing: 1px; text-transform: uppercase; margin-top: 4px; }

/* ── Toolbar ── */
.toolbar { display: flex; gap: 10px; margin-bottom: 20px; align-items: center; flex-wrap: wrap; }
.search-input {
  flex: 1; min-width: 200px;
  background: var(--surface); border: 1px solid var(--border);
  color: var(--text); padding: 8px 14px; border-radius: 6px;
  font-family: var(--sans); font-size: 13px; outline: none;
}
.search-input:focus { border-color: rgba(239,68,68,.4); }
.filter-select {
  background: var(--surface); border: 1px solid var(--border);
  color: var(--text); padding: 8px 12px; border-radius: 6px;
  font-family: var(--mono); font-size: 11px; outline: none; cursor: pointer;
}

/* ── Table ── */
.cve-table { width: 100%; border-collapse: collapse; }
.cve-table th {
  font-family: var(--mono); font-size: 9px; color: var(--muted);
  letter-spacing: 2px; text-transform: uppercase;
  padding: 10px 14px; border-bottom: 1px solid var(--border);
  text-align: left; white-space: nowrap; background: var(--surface);
}
.cve-table td { padding: 12px 14px; border-bottom: 1px solid rgba(30,41,59,.6); font-size: 13px; vertical-align: middle; }
.cve-table tr { cursor: pointer; transition: .1s; }
.cve-table tr:hover td { background: rgba(255,255,255,.02); }
.cve-table tr:last-child td { border-bottom: none; }

.sev-badge, .status-badge {
  display: inline-block; padding: 2px 8px; border-radius: 3px;
  font-family: var(--mono); font-size: 10px; letter-spacing: 1px;
  text-transform: uppercase; white-space: nowrap;
  border: 1px solid currentColor;
}
.cve-id { font-family: var(--mono); font-size: 12px; font-weight: 600; color: #60a5fa; }
.filter-name { font-size: 12px; color: var(--muted); max-width: 300px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.has-notes::after { content: '●'; color: #f97316; font-size: 8px; margin-left: 4px; vertical-align: middle; }

/* ── Detail panel ── */
.detail-overlay {
  position: fixed; inset: 0; background: rgba(0,0,0,.7); z-index: 100;
  display: none; align-items: center; justify-content: center; padding: 32px;
  backdrop-filter: blur(4px);
}
.detail-overlay.show { display: flex; }
.detail-panel {
  background: var(--surface); border: 1px solid var(--border); border-radius: 12px;
  width: 100%; max-width: 680px; max-height: 85vh; overflow-y: auto;
  padding: 32px; position: relative;
}
.detail-close {
  position: absolute; top: 16px; right: 16px;
  background: none; border: none; color: var(--muted);
  font-size: 18px; cursor: pointer; line-height: 1; padding: 4px 8px;
}
.detail-close:hover { color: var(--text); }
.detail-cve { font-family: var(--mono); font-size: 20px; font-weight: 700; color: #60a5fa; margin-bottom: 4px; }
.detail-filter { font-size: 14px; color: var(--muted); margin-bottom: 20px; line-height: 1.5; }
.detail-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-bottom: 20px; }
.detail-field { background: var(--surface2); border: 1px solid var(--border); border-radius: 6px; padding: 12px; }
.detail-field label { font-family: var(--mono); font-size: 9px; color: var(--muted); letter-spacing: 2px; text-transform: uppercase; display: block; margin-bottom: 5px; }
.detail-field p { font-size: 12px; line-height: 1.6; }
.detail-field.full { grid-column: 1 / -1; }
.status-select {
  width: 100%; background: var(--surface2); border: 1px solid var(--border);
  color: var(--text); padding: 8px 12px; border-radius: 6px;
  font-family: var(--mono); font-size: 12px; outline: none; margin-top: 4px;
}
.notes-textarea {
  width: 100%; background: var(--surface2); border: 1px solid var(--border);
  color: var(--text); padding: 10px 12px; border-radius: 6px;
  font-family: var(--sans); font-size: 13px; outline: none;
  resize: vertical; min-height: 80px; margin-top: 4px; line-height: 1.6;
}
.notes-textarea:focus, .status-select:focus { border-color: rgba(239,68,68,.4); }
.btn-save {
  background: #ef4444; color: #fff; border: none; padding: 9px 20px;
  border-radius: 6px; font-family: var(--mono); font-size: 11px;
  letter-spacing: 1px; cursor: pointer; margin-top: 12px;
}
.btn-save:hover { background: #dc2626; }
.saved-flash { color: #22c55e; font-family: var(--mono); font-size: 11px; margin-left: 10px; display: none; }

.page-title { font-size: 22px; font-weight: 600; margin-bottom: 4px; }
.page-sub { font-size: 13px; color: var(--muted); margin-bottom: 24px; }

/* scan bar at top */
.scan-bar {
  background: var(--surface); border: 1px solid var(--border); border-radius: 8px;
  padding: 12px 16px; margin-bottom: 20px;
  display: flex; align-items: center; gap: 12px; flex-wrap: wrap;
}
.scan-input {
  flex: 1; min-width: 200px;
  background: var(--bg); border: 1px solid var(--border);
  color: var(--text); padding: 7px 12px; border-radius: 5px;
  font-family: var(--mono); font-size: 11px; outline: none;
}
.btn-scan {
  background: rgba(239,68,68,.15); color: #ef4444;
  border: 1px solid rgba(239,68,68,.3); padding: 7px 16px;
  border-radius: 5px; font-family: var(--mono); font-size: 11px;
  letter-spacing: 1px; cursor: pointer; white-space: nowrap;
}
.btn-scan:hover { background: rgba(239,68,68,.25); }
.scan-status { font-family: var(--mono); font-size: 11px; color: var(--muted); }
</style>
</head>
<body>
<div class="shell">

<!-- Sidebar -->
<div class="sidebar">
  <div class="logo">CVE TRACKER<span>// NOT ENABLED BY DEFAULT</span></div>

  <div class="nav-label">Filter</div>
  <a class="nav-item {% if status_filter == 'all' %}active{% endif %}" href="/?status=all&sev={{ sev_filter }}&mode={{ mode }}">
    ◈ All CVEs <span class="nav-count">{{ counts.total }}</span>
  </a>
  <a class="nav-item {% if status_filter == 'open' %}active{% endif %}" href="/?status=open&sev={{ sev_filter }}&mode={{ mode }}">
    ⬤ Open <span class="nav-count">{{ counts.open }}</span>
  </a>
  <a class="nav-item {% if status_filter == 'investigating' %}active{% endif %}" href="/?status=investigating&sev={{ sev_filter }}&mode={{ mode }}">
    ◎ Investigating <span class="nav-count">{{ counts.investigating }}</span>
  </a>
  <a class="nav-item {% if status_filter == 'accepted' %}active{% endif %}" href="/?status=accepted&sev={{ sev_filter }}&mode={{ mode }}">
    ◇ Accepted <span class="nav-count">{{ counts.accepted }}</span>
  </a>
  <a class="nav-item {% if status_filter == 'resolved' %}active{% endif %}" href="/?status=resolved&sev={{ sev_filter }}&mode={{ mode }}">
    ✓ Resolved <span class="nav-count">{{ counts.resolved }}</span>
  </a>
  <a class="nav-item {% if status_filter == 'false_positive' %}active{% endif %}" href="/?status=false_positive&sev={{ sev_filter }}&mode={{ mode }}">
    ✗ False Positive <span class="nav-count">{{ counts.false_positive }}</span>
  </a>

  <div style="margin-top:20px;">
  <div class="nav-label">Severity</div>
  {% for sev, color in sev_colors.items() %}
  <a class="nav-item {% if sev_filter == sev %}active{% endif %}" href="/?sev={{ sev }}&status={{ status_filter }}&mode={{ mode }}" style="border-left-color: {{ color if sev_filter == sev else 'transparent' }}">
    <span style="color:{{ color }}">■</span> {{ sev }}
    <span class="nav-count">{{ counts.get('sev_' + sev, 0) }}</span>
  </a>
  {% endfor %}
  </div>

  <div class="sidebar-footer">{{ db_path }}</div>
</div>

<!-- Main -->
<div class="main">
  <div class="page-title">CVE Intelligence</div>
  <div class="page-sub">TrendAI DV Filters — Weekly Release Review</div>

  <!-- Mode toggle -->
  <div class="mode-toggle">
    <a class="mode-btn {% if mode == 'not_enabled' %}active{% endif %}"
       href="/?mode=not_enabled&status={{ status_filter }}&sev={{ sev_filter }}">
      Not Enabled by Default
    </a>
    <a class="mode-btn {% if mode == 'block_notify' %}active{% endif %}"
       href="/?mode=block_notify&status={{ status_filter }}&sev={{ sev_filter }}">
      Block &amp; Notify
    </a>
    <a class="mode-btn {% if mode == 'all' %}active{% endif %}"
       href="/?mode=all&status={{ status_filter }}&sev={{ sev_filter }}">
      All Filters
    </a>
  </div>

  <!-- Scan bar -->
  <div class="scan-bar">
    <span style="font-family:var(--mono);font-size:10px;color:var(--muted);letter-spacing:1px;">SCAN:</span>
    <input class="scan-input" id="scan-path" value="{{ default_dir }}" placeholder="/path/to/DVSheets">
    <button class="btn-scan" onclick="runScan()">▶ Scan Now</button>
    <span class="scan-status" id="scan-status"></span>
  </div>

  <!-- Stats -->
  <div class="stats-bar">
    <div class="stat"><div class="stat-num" style="color:#ef4444">{{ counts.get('sev_Critical', 0) }}</div><div class="stat-label">Critical</div></div>
    <div class="stat"><div class="stat-num" style="color:#f97316">{{ counts.get('sev_High', 0) }}</div><div class="stat-label">High</div></div>
    <div class="stat"><div class="stat-num" style="color:#eab308">{{ counts.get('sev_Medium', 0) + counts.get('sev_Moderate', 0) }}</div><div class="stat-label">Medium</div></div>
    <div class="stat"><div class="stat-num" style="color:#ef4444">{{ counts.open }}</div><div class="stat-label">Open</div></div>
    <div class="stat"><div class="stat-num" style="color:#22c55e">{{ counts.resolved }}</div><div class="stat-label">Resolved</div></div>
    <div class="stat"><div class="stat-num">{{ counts.total }}</div><div class="stat-label">Total</div></div>
  </div>

  <!-- Toolbar -->
  <div class="toolbar">
    <input class="search-input" type="text" id="search" placeholder="Search CVE ID, filter name, affected system..." oninput="filterRows()">
    <select class="filter-select" id="sev-select" onchange="filterRows()">
      <option value="">All Severities</option>
      {% for sev in ['Critical','High','Medium','Moderate','Low','Informational'] %}
      <option value="{{ sev }}" {% if sev_filter == sev %}selected{% endif %}>{{ sev }}</option>
      {% endfor %}
    </select>
  </div>

  <!-- Table -->
  <table class="cve-table" id="cve-table">
    <thead>
      <tr>
        <th>CVE ID</th>
        <th>Severity</th>
        <th>Status</th>
        <th>Filter</th>
        <th>Enabled Setting</th>
        <th>Affected</th>
        <th>First Seen</th>
        <th>Last Seen</th>
      </tr>
    </thead>
    <tbody>
      {% for row in cves %}
      <tr onclick="openDetail({{ row.id }})"
          data-text="{{ (row.cve_id + ' ' + row.filter_desc + ' ' + row.affected + ' ' + row.severity)|lower }}"
          data-sev="{{ row.severity }}">
        <td>
          <span class="cve-id {% if row.notes %}has-notes{% endif %}">{{ row.cve_id }}</span>
        </td>
        <td>
          <span class="sev-badge" style="color:{{ sev_colors.get(row.severity, '#64748b') }};border-color:{{ sev_colors.get(row.severity, '#64748b') }}20;background:{{ sev_colors.get(row.severity, '#64748b') }}12;">
            {{ row.severity }}
          </span>
        </td>
        <td>
          <span class="status-badge" style="color:{{ status_colors.get(row.status, '#64748b') }};border-color:{{ status_colors.get(row.status,'#64748b') }}30;background:{{ status_colors.get(row.status,'#64748b') }}10;">
            {{ row.status.replace('_',' ') }}
          </span>
        </td>
        <td><span class="filter-name">{{ row.filter_desc }}</span></td>
        <td style="font-size:11px;color:var(--muted);font-family:var(--mono);max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">
          {% if 'Not enabled' in (row.enabled_recommendation or '') %}
            <span style="color:#ef4444">✗ Not Enabled</span>
          {% elif 'Block' in (row.enabled_recommendation or '') %}
            <span style="color:#22c55e">✓ Block/Notify</span>
          {% else %}
            {{ (row.enabled_recommendation or '—')[:20] }}
          {% endif %}
        </td>
        <td style="font-size:12px;color:var(--muted);">{{ row.affected[:40] }}{% if row.affected|length > 40 %}…{% endif %}</td>
        <td style="font-family:var(--mono);font-size:11px;color:var(--muted);">{{ row.first_seen[:10] }}</td>
        <td style="font-family:var(--mono);font-size:11px;color:var(--muted);">{{ row.last_seen[:10] }}</td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
</div>

<!-- Detail overlay -->
<div class="detail-overlay" id="detail-overlay" onclick="closeDetailOutside(event)">
  <div class="detail-panel" id="detail-panel">
    <button class="detail-close" onclick="closeDetail()">✕</button>
    <div class="detail-cve" id="d-cve"></div>
    <div class="detail-filter" id="d-filter"></div>
    <div class="detail-grid">
      <div class="detail-field">
        <label>Severity</label>
        <p id="d-severity"></p>
      </div>
      <div class="detail-field">
        <label>Category</label>
        <p id="d-category"></p>
      </div>
      <div class="detail-field full">
        <label>Affected Systems</label>
        <p id="d-affected"></p>
      </div>
      <div class="detail-field full">
        <label>Potential False Positive</label>
        <p id="d-false-pos"></p>
      </div>
      <div class="detail-field full">
        <label>Potential Performance Impact</label>
        <p id="d-perf"></p>
      </div>
      <div class="detail-field">
        <label>Source File</label>
        <p id="d-source" style="font-family:var(--mono);font-size:11px;"></p>
      </div>
      <div class="detail-field">
        <label>First Seen / Last Seen</label>
        <p id="d-dates" style="font-family:var(--mono);font-size:11px;"></p>
      </div>
      <div class="detail-field full">
        <label>Status</label>
        <select class="status-select" id="d-status">
          <option value="open">Open</option>
          <option value="investigating">Investigating</option>
          <option value="accepted">Accepted</option>
          <option value="resolved">Resolved</option>
          <option value="false_positive">False Positive</option>
        </select>
      </div>
      <div class="detail-field full">
        <label>Notes</label>
        <textarea class="notes-textarea" id="d-notes" placeholder="Add investigation notes, resolution details..."></textarea>
      </div>
    </div>
    <button class="btn-save" onclick="saveDetail()">Save Changes</button>
    <span class="saved-flash" id="saved-flash">✓ Saved</span>
  </div>
</div>

<script>
let currentId = null;

function filterRows() {
  const q = document.getElementById('search').value.toLowerCase();
  const sev = document.getElementById('sev-select').value.toLowerCase();
  document.querySelectorAll('#cve-table tbody tr').forEach(tr => {
    const text = tr.dataset.text || '';
    const rowSev = (tr.dataset.sev || '').toLowerCase();
    const matchQ = !q || text.includes(q);
    const matchS = !sev || rowSev === sev;
    tr.style.display = (matchQ && matchS) ? '' : 'none';
  });
}

function openDetail(id) {
  fetch('/api/cve/' + id)
    .then(r => r.json())
    .then(d => {
      currentId = id;
      document.getElementById('d-cve').textContent = d.cve_id;
      document.getElementById('d-filter').textContent = d.filter_desc;
      document.getElementById('d-severity').textContent = d.severity;
      document.getElementById('d-category').textContent = d.category;
      document.getElementById('d-affected').textContent = d.affected;
      document.getElementById('d-false-pos').textContent = d.false_positive || '—';
      document.getElementById('d-perf').textContent = d.performance || '—';
      document.getElementById('d-source').textContent = d.sheet + ' / ' + d.source_file;
      document.getElementById('d-dates').textContent = (d.first_seen||'').slice(0,10) + '  →  ' + (d.last_seen||'').slice(0,10);
      document.getElementById('d-status').value = d.status;
      document.getElementById('d-notes').value = d.notes || '';
      document.getElementById('saved-flash').style.display = 'none';
      document.getElementById('detail-overlay').classList.add('show');
      document.body.style.overflow = 'hidden';
    });
}

function closeDetail() {
  document.getElementById('detail-overlay').classList.remove('show');
  document.body.style.overflow = '';
  currentId = null;
}

function closeDetailOutside(e) {
  if (e.target === document.getElementById('detail-overlay')) closeDetail();
}

function saveDetail() {
  if (!currentId) return;
  const status = document.getElementById('d-status').value;
  const notes  = document.getElementById('d-notes').value;
  fetch('/api/cve/' + currentId, {
    method: 'POST',
    headers: {'Content-Type':'application/json'},
    body: JSON.stringify({status, notes})
  }).then(r => r.json()).then(d => {
    if (d.ok) {
      const flash = document.getElementById('saved-flash');
      flash.style.display = 'inline';
      setTimeout(() => flash.style.display = 'none', 2000);
      // Update badge in table
      location.reload();
    }
  });
}

function runScan() {
  const path = document.getElementById('scan-path').value;
  const statusEl = document.getElementById('scan-status');
  statusEl.textContent = 'Scanning...';
  fetch('/api/scan', {
    method: 'POST',
    headers: {'Content-Type':'application/json'},
    body: JSON.stringify({path})
  }).then(r => r.json()).then(d => {
    statusEl.textContent = d.message;
    setTimeout(() => location.reload(), 1500);
  }).catch(() => { statusEl.textContent = 'Error'; });
}

document.addEventListener('keydown', e => { if (e.key === 'Escape') closeDetail(); });
</script>
</body>
</html>
"""

# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    conn = get_db()
    status_filter = request.args.get('status', 'all')
    sev_filter = request.args.get('sev', '')
    mode = request.args.get('mode', 'not_enabled')

    query = 'SELECT * FROM cves'
    params = []
    conditions = []
    if mode == 'not_enabled':
        conditions.append('enabled_recommendation=?')
        params.append('Not enabled by default in any deployment')
    elif mode == 'block_notify':
        conditions.append("(enabled_recommendation LIKE '%Block / Notify%')")
    if status_filter != 'all' and status_filter in VALID_STATUSES:
        conditions.append('status=?')
        params.append(status_filter)
    if sev_filter:
        conditions.append('severity=?')
        params.append(sev_filter)
    if conditions:
        query += ' WHERE ' + ' AND '.join(conditions)
    query += ''' ORDER BY
        CASE severity WHEN 'Critical' THEN 0 WHEN 'High' THEN 1
        WHEN 'Medium' THEN 2 WHEN 'Moderate' THEN 2 WHEN 'Low' THEN 3 ELSE 4 END,
        cve_id'''

    cves = conn.execute(query, params).fetchall()

    # Counts
    all_rows = conn.execute('SELECT status, severity FROM cves').fetchall()
    counts = {'total': len(all_rows), 'open': 0, 'investigating': 0,
              'accepted': 0, 'resolved': 0, 'false_positive': 0}
    for r in all_rows:
        if r['status'] in counts:
            counts[r['status']] += 1
        sev_key = 'sev_' + r['severity']
        counts[sev_key] = counts.get(sev_key, 0) + 1

    conn.close()

    return render_template_string(HTML,
        cves=cves,
        counts=counts,
        status_filter=status_filter,
        mode=mode,
        sev_filter=sev_filter,
        sev_colors=SEVERITY_COLORS,
        status_colors=STATUS_COLORS,
        db_path=os.path.basename(DB_FILE),
        default_dir='/Users/stephenb/Documents/Trend/TippingPoint/DVSheets',
    )

@app.route('/api/cve/<int:cve_id>')
def api_get_cve(cve_id):
    conn = get_db()
    row = conn.execute('SELECT * FROM cves WHERE id=?', (cve_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'not found'}), 404
    return jsonify(dict(row))

@app.route('/api/cve/<int:cve_id>', methods=['POST'])
def api_update_cve(cve_id):
    data = request.get_json()
    status = data.get('status')
    notes  = data.get('notes', '')
    if status not in VALID_STATUSES:
        return jsonify({'ok': False, 'error': 'invalid status'}), 400
    conn = get_db()
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    conn.execute('UPDATE cves SET status=?, notes=?, updated_at=? WHERE id=?',
                 (status, notes, now, cve_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/scan', methods=['POST'])
def api_scan():
    import glob as g
    from openpyxl import load_workbook

    data = request.get_json()
    path = data.get('path', '')

    if os.path.isfile(path):
        files = [path]
    elif os.path.isdir(path):
        files = sorted(g.glob(os.path.join(path, '*.xlsx')) +
                       g.glob(os.path.join(path, '*.xlsm')))
    else:
        return jsonify({'message': f'Path not found: {path}'})

    if not files:
        return jsonify({'message': 'No .xlsx files found'})

    conn = get_db()
    new_count = updated_count = 0
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    for filepath in files:
        try:
            wb = load_workbook(filepath, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) <= 9: continue
                    enabled = str(row[9]).strip() if row[9] else ''
                    if not enabled: continue
                    severity = str(row[3]).strip() if row[3] else ''
                    cve_raw  = str(row[5]).strip() if row[5] else ''
                    cve_ids  = [p.strip() for p in cve_raw.replace('\n',' ').split()
                                if p.strip().upper().startswith('CVE-')]
                    cve_str  = ', '.join(cve_ids) if cve_ids else cve_raw
                    fp       = str(row[10]).strip() if len(row)>10 and row[10] else ''
                    perf     = str(row[11]).strip() if len(row)>11 and row[11] else ''

                    existing = conn.execute('SELECT id FROM cves WHERE cve_id=?', (cve_str,)).fetchone()
                    if existing:
                        conn.execute('''UPDATE cves SET filter_desc=?,category=?,severity=?,
                            affected=?,false_positive=?,performance=?,sheet=?,source_file=?,
                            enabled_recommendation=?,last_seen=?,updated_at=? WHERE cve_id=?''',
                            (str(row[1]).strip() if row[1] else '',
                             str(row[2]).strip() if row[2] else '',
                             severity, str(row[4]).strip() if row[4] else '',
                             fp, perf, sheet_name, os.path.basename(filepath),
                             enabled, now, now, cve_str))
                        updated_count += 1
                    else:
                        conn.execute('''INSERT INTO cves
                            (cve_id,filter_desc,category,severity,affected,false_positive,
                             performance,sheet,source_file,first_seen,last_seen,status,notes,
                             enabled_recommendation,updated_at)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                            (cve_str, str(row[1]).strip() if row[1] else '',
                             str(row[2]).strip() if row[2] else '',
                             severity, str(row[4]).strip() if row[4] else '',
                             fp, perf, sheet_name, os.path.basename(filepath),
                             now, now, 'open', '', enabled, now))
                        new_count += 1
            wb.close()
        except Exception as e:
            conn.close()
            return jsonify({'message': f'Error: {e}'})

    conn.commit()
    conn.close()
    return jsonify({'message': f'{new_count} new, {updated_count} updated — reloading...'})


if __name__ == '__main__':
    print(f"\nCVE Tracker running at http://localhost:5001")
    print(f"Database: {DB_FILE}\n")
    app.run(port=5001, debug=False)
