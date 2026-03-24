#!/usr/bin/env python3
"""
Extract Critical CVEs marked "Not enabled by default in any deployment"
from TrendAI Weekly DV Release List spreadsheets.

Usage:
    python3 extract_critical_cves.py /path/to/directory
    python3 extract_critical_cves.py /path/to/single_file.xlsx

Output:
    - Prints a summary to the console
    - Saves extract_critical_cves.csv in the current directory
"""

import sys
import os
import csv
import glob
from openpyxl import load_workbook

SEVERITY_COL    = 3   # TrendAI Defined Severity
CVE_COL         = 5   # CVE ID + CVSS 3.0
ENABLED_COL     = 9   # Enabled TrendAI Recommendations
FILTER_DESC_COL = 1   # TrendAI Filter & Description
CATEGORY_COL    = 2   # Category
AFFECTED_COL    = 4   # Affected Systems

TARGET_SEVERITY = 'Critical'
TARGET_ENABLED  = 'Not enabled by default in any deployment'

# Check all sheets
SKIP_SHEETS = ['Removed Filters']


def extract_from_sheet(ws, filename, sheet_name):
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= ENABLED_COL:
            continue
        severity = str(row[SEVERITY_COL]).strip() if row[SEVERITY_COL] else ''
        enabled  = str(row[ENABLED_COL]).strip() if row[ENABLED_COL] else ''

        if severity == TARGET_SEVERITY and enabled == TARGET_ENABLED:
            cve_raw = str(row[CVE_COL]).strip() if row[CVE_COL] else ''
            cve_ids = [p.strip() for p in cve_raw.replace('\n', ' ').split()
                       if p.strip().upper().startswith('CVE-')]
            cve_str = ', '.join(cve_ids) if cve_ids else cve_raw

            results.append({
                'File':      os.path.basename(filename),
                'Sheet':     sheet_name,
                'CVE ID(s)': cve_str,
                'Filter':    str(row[FILTER_DESC_COL]).strip() if row[FILTER_DESC_COL] else '',
                'Category':  str(row[CATEGORY_COL]).strip() if row[CATEGORY_COL] else '',
                'Severity':  severity,
                'Enabled':   enabled,
                'Affected':  str(row[AFFECTED_COL]).strip() if row[AFFECTED_COL] else '',
            })
    return results


def process_file(filepath):
    results = []
    try:
        wb = load_workbook(filepath, read_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            ws = wb[sheet_name]
            found = extract_from_sheet(ws, filepath, sheet_name)
            results.extend(found)
        wb.close()
    except Exception as e:
        print(f"  ERROR reading {filepath}: {e}")
    return results


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 extract_critical_cves.py <directory_or_file>")
        sys.exit(1)

    target = sys.argv[1]

    if os.path.isfile(target):
        files = [target]
    elif os.path.isdir(target):
        files = sorted(
            glob.glob(os.path.join(target, '*.xlsx')) +
            glob.glob(os.path.join(target, '*.xlsm'))
        )
        if not files:
            print(f"No .xlsx files found in {target}")
            sys.exit(1)
    else:
        print(f"Not a valid file or directory: {target}")
        sys.exit(1)

    print(f"\nProcessing {len(files)} file(s)...\n")

    all_results = []
    for f in files:
        print(f"  → {os.path.basename(f)}")
        rows = process_file(f)
        print(f"     Found {len(rows)} Critical + Not Enabled row(s)")
        all_results.extend(rows)

    print(f"\n{'='*60}")
    print(f"TOTAL: {len(all_results)} Critical CVE(s) not enabled by default")
    print(f"{'='*60}\n")

    if all_results:
        for r in all_results:
            print(f"  CVE:     {r['CVE ID(s)']}")
            print(f"  Filter:  {r['Filter']}")
            print(f"  Sheet:   {r['Sheet']} in {r['File']}")
            print()

        out_file = 'extract_critical_cves.csv'
        with open(out_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=[
                'File', 'Sheet', 'CVE ID(s)', 'Filter', 'Category',
                'Severity', 'Enabled', 'Affected'
            ])
            writer.writeheader()
            writer.writerows(all_results)
        print(f"Saved: {out_file}")
    else:
        print("No Critical CVEs marked 'Not enabled by default' found in these files.")
        print("(This is expected for weeks where all Critical filters are enabled by default.)")


if __name__ == '__main__':
    main()